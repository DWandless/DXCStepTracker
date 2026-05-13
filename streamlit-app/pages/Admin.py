import streamlit as st
import os
import shutil
import zipfile
import io
import pandas as pd
import re
import unicodedata
import time
import json
from datetime import datetime
from pathlib import Path
from db import supabase
from components import (apply_dxc_theme, setup_logo, render_header, render_footer, render_sidebar_welcome,
                        hide_streamlit_branding, check_login_required, handle_logout, secure_filename, log_audit_event)
from onedrive_storage import get_file_download_url, delete_from_onedrive, get_access_token, get_file_id_from_sharing_url
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------ PAGE CONFIG ------------------
logo_path = Path(__file__).resolve().parents[1] / ".streamlit" / "static" / "assets" / "logo.png"
st.set_page_config(page_title="Admin Dashboard", layout="wide", page_icon=logo_path)

# Hide branding early
hide_streamlit_branding()

# ------------------ APPLY THEME & LOGO ------------------
apply_dxc_theme()
setup_logo()
render_header("Admin Dashboard", "Manage submissions and verify evidence.")

# ------------------ LOGIN & ROLE CHECK ------------------
username = check_login_required()
user_email = st.session_state.get("user_email", "")

# Check if user email is in admin list from secrets
admin_emails = st.secrets.get("ADMIN_EMAILS", [])
if user_email.lower() not in [email.lower() for email in admin_emails]:
    log_audit_event("ADMIN_ACCESS_DENIED", user_email, "Attempted to access admin dashboard")
    st.error("Access denied.")
    st.stop()

# ------------------ SECURITY FUNCTION ------------------
# secure_filename now imported from components

# ------------------ CONFIG & STATE ------------------
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# replace old confirm state with a simpler pending_delete entry
if "pending_delete" not in st.session_state:
    st.session_state["pending_delete"] = None

# ------------------ FETCH DATA FROM SUPABASE ------------------
def fetch_all_submissions():
    forms = supabase.table("forms") \
        .select("*") \
        .eq("form_verified", False) \
        .gt("form_stepcount", 19999) \
        .execute().data
    users = supabase.table("users").select("user_id, user_name").execute().data
    if not forms:
        return pd.DataFrame()
    df_forms = pd.DataFrame(forms)
    df_users = pd.DataFrame(users)
    return pd.merge(df_forms, df_users, on="user_id")

df = fetch_all_submissions()

# ------------------ GENERATE COMPREHENSIVE EXCEL EXPORT ------------------
def generate_comprehensive_export():
    """Generate comprehensive Excel export with user and team views."""
    try:
        # Fetch all users with their forms
        users = supabase.table("users").select("*").execute().data
        forms = supabase.table("forms").select("*").execute().data
        teams = supabase.table("teams").select("*").execute().data
        
        # Load challenges data
        challenges_path = Path(__file__).resolve().parents[1] / ".streamlit" / "static" / "assets" / "Challenges.json"
        with open(challenges_path, "r") as f:
            challenges_data = json.load(f)
        
        # Create workbook
        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "Users"
        ws_teams = wb.create_sheet("Teams")
        
        # Define DXC blue color
        dxc_blue = "7BA4DB"
        header_fill = PatternFill(start_color=dxc_blue, end_color=dxc_blue, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== USERS SHEET ====================
        user_headers = [
            "Name", "Email", "Team Name", "Team Leader",
            "Total Submissions", "Total Steps", "Average Daily Steps",
            "Max Steps (Single Day)", "First Submission Date", "Last Submission Date",
            "Challenges Completed"
        ]
        
        # Write headers
        for col_num, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Calculate user statistics
        user_stats = []
        for user in users:
            user_forms = [f for f in forms if f["user_id"] == user["user_id"]]
            total_submissions = len(user_forms)
            total_steps = sum(f["form_stepcount"] for f in user_forms)
            avg_daily = total_steps / total_submissions if total_submissions > 0 else 0
            max_steps = max((f["form_stepcount"] for f in user_forms), default=0)
            
            dates = [f["form_date"] for f in user_forms if f["form_date"]]
            first_date = min(dates) if dates else "N/A"
            last_date = max(dates) if dates else "N/A"
            
            # Get team info
            team_name = "No Team"
            team_leader = "N/A"
            if user.get("team_id"):
                team = next((t for t in teams if t["team_id"] == user["team_id"]), None)
                if team:
                    team_name = team["team_name"]
                    # Get team leader name
                    leader = next((u for u in users if u["user_id"] == team["team_leader_id"]), None)
                    team_leader = leader["user_name"] if leader else "N/A"
            
            # Get completed challenges
            completed_challenges = []
            for challenge_key, challenge_data in challenges_data.items():
                if challenge_key in user_forms:
                    completed_challenges.append(challenge_data["title"])
            challenges_str = ", ".join(completed_challenges) if completed_challenges else "None"
            
            user_stats.append({
                "Name": user["user_name"],
                "Email": user["user_email"],
                "Team Name": team_name,
                "Team Leader": team_leader,
                "Total Submissions": total_submissions,
                "Total Steps": total_steps,
                "Average Daily Steps": round(avg_daily, 2),
                "Max Steps (Single Day)": max_steps,
                "First Submission Date": first_date,
                "Last Submission Date": last_date,
                "Challenges Completed": challenges_str
            })
        
        # Sort by total steps (descending)
        user_stats.sort(key=lambda x: x["Total Steps"], reverse=True)
        
        # Write user data
        for row_num, stat in enumerate(user_stats, 2):
            for col_num, header in enumerate(user_headers, 1):
                cell = ws_users.cell(row=row_num, column=col_num)
                cell.value = stat[header]
                cell.border = border_style
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for col_num, header in enumerate(user_headers, 1):
            max_length = len(header)
            for row_num in range(2, len(user_stats) + 2):
                cell_value = ws_users.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_users.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # ==================== TEAMS SHEET ====================
        team_headers = [
            "Team Name", "Leader", "Member Count", "Total Team Steps",
            "Average Steps per Member", "Total Submissions", "Average Daily Steps per Member",
            "Max Steps (Single Day)", "Members"
        ]
        
        # Write headers
        for col_num, header in enumerate(team_headers, 1):
            cell = ws_teams.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Calculate team statistics
        team_stats = []
        for team in teams:
            team_users = [u for u in users if u.get("team_id") == team["team_id"]]
            member_count = len(team_users)
            
            # Get team leader name
            leader = next((u for u in users if u["user_id"] == team["team_leader_id"]), None)
            leader_name = leader["user_name"] if leader else "N/A"
            
            # Calculate team statistics
            team_forms = [f for f in forms if f["user_id"] in [u["user_id"] for u in team_users]]
            total_team_steps = sum(f["form_stepcount"] for f in team_forms)
            avg_steps_per_member = total_team_steps / member_count if member_count > 0 else 0
            total_submissions = len(team_forms)
            avg_daily_per_member = total_team_steps / member_count if member_count > 0 else 0
            max_steps = max((f["form_stepcount"] for f in team_forms), default=0)
            
            # Get member names
            member_names = ", ".join([u["user_name"] for u in team_users])
            
            team_stats.append({
                "Team Name": team["team_name"],
                "Leader": leader_name,
                "Member Count": member_count,
                "Total Team Steps": total_team_steps,
                "Average Steps per Member": round(avg_steps_per_member, 2),
                "Total Submissions": total_submissions,
                "Average Daily Steps per Member": round(avg_daily_per_member, 2),
                "Max Steps (Single Day)": max_steps,
                "Members": member_names
            })
        
        # Sort by total team steps (descending)
        team_stats.sort(key=lambda x: x["Total Team Steps"], reverse=True)
        
        # Write team data
        for row_num, stat in enumerate(team_stats, 2):
            for col_num, header in enumerate(team_headers, 1):
                cell = ws_teams.cell(row=row_num, column=col_num)
                cell.value = stat[header]
                cell.border = border_style
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for col_num, header in enumerate(team_headers, 1):
            max_length = len(header)
            for row_num in range(2, len(team_stats) + 2):
                cell_value = ws_teams.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_teams.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    except Exception as e:
        logging.error(f"Error generating comprehensive export: {e}")
        return None

# ------------------ SIDEBAR ------------------
if render_sidebar_welcome():
    handle_logout()

# ------------------ 1. HIGH-STEP SUBMISSIONS (>20,000) ------------------
st.subheader(
    "Unverified Submissions (Steps > 20,000)",
    help="Review and verify step submissions over 20,000 steps. Check the screenshot evidence and click 'Verify' to approve or 'Delete' to remove suspicious entries."
)

if not df.empty:
    for idx, row in df.iterrows():
        col1, col2 = st.columns([3, 1])
        filepath = row.get("form_filepath", "")
        
        # Check if it's a OneDrive URL or local file
        is_onedrive = filepath.startswith("http") or "sharepoint" in filepath.lower() or "onedrive" in filepath.lower()
        
        with col1:
            st.markdown(f"**Name:** {row['user_name']} | **Date:** {row['form_date']} | **Steps:** {row['form_stepcount']}")
            with st.expander("View Full Screenshot"):
                if is_onedrive:
                    st.markdown(f"**Evidence stored in OneDrive:**" + f" [🗁 Open in OneDrive]({filepath})")
                    st.info("Click the link above to view the screenshot in OneDrive.")
                else:
                    safe_name = secure_filename(os.path.basename(str(filepath)))
                    file_path = os.path.join(UPLOAD_FOLDER, safe_name)
                    if os.path.exists(file_path):
                        st.image(file_path, caption=f"Screenshot for {row['user_name']}", width="stretch")
                    else:
                        st.warning("Screenshot not found.")

        with col2:
            if st.button("Verify", key=f"verify_{idx}"):
                try:
                    # Delete from OneDrive if applicable
                    if is_onedrive:
                        access_token = get_access_token()
                        if access_token:
                            file_id = get_file_id_from_sharing_url(filepath, access_token)
                            if file_id:
                                delete_from_onedrive(file_id, access_token)
                    
                    # Delete local file if applicable
                    if not is_onedrive:
                        try:
                            safe_name = secure_filename(os.path.basename(str(filepath)))
                            file_path = os.path.join(UPLOAD_FOLDER, safe_name)
                            os.remove(file_path)
                        except FileNotFoundError:
                            pass

                    # Update database to set form_verified to True
                    supabase.table("forms") \
                        .update({"form_verified": True}) \
                        .eq("form_id", row["form_id"]) \
                        .execute()

                    log_audit_event("VERIFICATION", user_email, f"Form ID: {row['form_id']}, Steps: {row['form_stepcount']}")
                    st.success("Submission verified successfully!")
                except Exception as e:
                    logging.error(f"Verification error: {str(e)}")
                    st.error("Error verifying form, please try again later.")
                st.rerun()
            if st.button("Delete", key=f"delete_{idx}"):
                # set a small pending_delete dict rather than relying on index
                st.session_state["pending_delete"] = {
                    "form_id": row["form_id"],
                    "user_name": row["user_name"],
                    "form_date": row["form_date"],
                    "form_stepcount": row["form_stepcount"],
                    "file": row.get("form_filepath", "")
                }
                st.rerun()
        
        # Show confirmation dialog inline if this row is being deleted
        if st.session_state["pending_delete"] and st.session_state["pending_delete"]["form_id"] == row["form_id"]:
            pd = st.session_state["pending_delete"]
            st.error("### ⚠ Confirm Deletion")
            st.markdown(f"You are about to permanently delete the submission for:")
            st.markdown(f"- **User:** {pd['user_name']}")
            st.markdown(f"- **Date:** {pd['form_date']}")
            st.markdown(f"- **Steps:** {pd['form_stepcount']}")
            st.markdown("")
            confirm_cb = st.checkbox("I understand this action cannot be undone", key="confirm_delete_cb")
            st.markdown("")
            colA, colB, colC = st.columns([1, 1, 2])
            with colA:
                if st.button("🗙 Delete Permanently", disabled=not confirm_cb, type="secondary", key=f"confirm_delete_{idx}"):
                    try:
                        filepath = pd.get("file", "")
                        is_onedrive = filepath.startswith("http") or "sharepoint" in filepath.lower() or "onedrive" in filepath.lower()
                        
                        # Transaction pattern: Delete from database first, then file
                        # If file deletion fails, we can't rollback, but at least we have the record
                        form_id = pd["form_id"]
                        file_path = pd.get("file", "")
                        
                        # Delete from database
                        supabase.table("forms").delete().eq("form_id", form_id).execute()
                        
                        # Delete file (local or OneDrive)
                        file_deleted = False
                        try:
                            if not is_onedrive:
                                safe_file_path = os.path.join(UPLOAD_FOLDER, secure_filename(os.path.basename(str(file_path))))
                                if os.path.exists(safe_file_path):
                                    os.remove(safe_file_path)
                                    file_deleted = True
                            else:
                                # Delete from OneDrive
                                access_token = get_access_token()
                                if access_token:
                                    file_id = get_file_id_from_sharing_url(file_path, access_token)
                                    if file_id:
                                        delete_from_onedrive(file_id, access_token)
                                        file_deleted = True
                        except Exception as file_error:
                            logging.error(f"File deletion failed but DB record deleted: {file_error}")
                            # File deletion failed but DB record is deleted
                            # This is acceptable - file will be orphaned but not accessible through app
                        
                        st.session_state["pending_delete"] = None
                        
                        if file_deleted:
                            st.success("Submission deleted successfully.")
                        else:
                            st.warning("Submission deleted from database, but file cleanup failed. This will be cleaned up later.")
                        
                        log_audit_event("DELETION", user_email, f"Form ID: {form_id}, File deleted: {file_deleted}")
                        st.rerun()
                    except Exception as e:
                        logging.error(f"Deletion error: {str(e)}")
                        st.error("Error deleting submission, please try again later.")
            with colB:
                if st.button("Cancel", key=f"cancel_delete_{idx}", type="secondary"):
                    st.session_state["pending_delete"] = None
                    st.rerun()
        
        st.markdown("---")
else:
    st.info("No high-step unverified submissions found.")

# ------------------ 2. DOWNLOAD COMPREHENSIVE DATA ------------------
st.subheader(
    "Download Comprehensive Data",
    help="Export comprehensive user and team statistics as an Excel file with detailed metrics, sorted by total steps."
)

excel_data = generate_comprehensive_export()
if excel_data:
    st.download_button(
        "Download Excel Report",
        excel_data,
        file_name=f"step_tracker_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.caption("Includes user statistics, team statistics, and challenge completion data.")
else:
    st.info("Unable to generate report. Please check the data and try again.")

st.divider()

# ------------------ 3. GENERATE CLAIM CODES -------------------
st.subheader(
    "Generate Claim Codes",
    help="Create unique claim codes for verified step submissions."
)
from components import get_all_challenges, generate_claim_code, hash_claim_code

# Initialize session state for generated codes
if "generated_codes" not in st.session_state:
    st.session_state.generated_codes = []

Challenges = get_all_challenges()

col_left, col_right = st.columns([1, 1])

with col_left:
    ChallengesDropdown = st.selectbox("Select Challenge", options=[Challenges[ch]["title"] for ch in Challenges])
    num_codes = st.number_input("Number of Claim Codes to Generate", min_value=1, max_value=100, value=5)
    
    if st.button("Generate Claim Codes"):
        if not ChallengesDropdown:
            st.error("Please select a challenge to generate claim codes for.")
            st.stop()

        # Generate codes with proper duplicate tracking
        generated_codes = []
        existing_codes = set()
        for challenge in Challenges:
            existing_codes.update(Challenges[challenge]["Codes"])
        
        for _ in range(num_codes):
            code = generate_claim_code(Challenges, existing_codes)
            generated_codes.append(code)
            existing_codes.add(hash_claim_code(code))  # Add hash to avoid duplicates

        # Hash codes for storage
        hashed_codes = [hash_claim_code(code) for code in generated_codes]

        # Read and update Challenges.json with proper path
        challenges_path = Path(__file__).resolve().parents[1] / ".streamlit" / "static" / "assets" / "Challenges.json"
        try:
            with open(challenges_path, "r") as f:
                challenges_data = json.load(f)
            
            # Add hashed codes to the selected challenge
            for challenge_key in challenges_data:
                if challenges_data[challenge_key]["title"] == ChallengesDropdown:
                    challenges_data[challenge_key]["Codes"].extend(hashed_codes)
                    break
            
            # Write back to file
            with open(challenges_path, "w") as f:
                json.dump(challenges_data, f, indent=4)

            st.session_state.generated_codes = generated_codes
            log_audit_event("CODE_GENERATION", user_email, f"Challenge: {ChallengesDropdown}, Count: {num_codes}")
        except Exception as e:
            st.error(f"Error generating codes. Please try again.")

with col_right:
    if st.session_state.generated_codes:
        codes_text = "\n".join(st.session_state.generated_codes)
        st.text_area("Codes", codes_text, height=200, key="generated_codes_display")
        
        # Download option
        st.download_button(
            label="Download Codes (Plain Text)",
            data=codes_text,
            file_name=f"claim_codes_{ChallengesDropdown.replace(' ', '_').lower()}.txt",
            mime="text/plain"
        )

# ------------------ FOOTER (ALWAYS RENDER) ------------------
render_footer()
